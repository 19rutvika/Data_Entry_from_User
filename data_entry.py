from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib

root=Tk()
root.title("Data Entry")
root.geometry('750x450+300+200')
root.resizable(False,False)
root.configure(bg="lightblue")

file=pathlib.Path("Backened_data.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Full Name"
    sheet['B1']="PhoneNumber"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"

    file.save("Backened_data.xlsx")
    


def submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=AgeValue.get()
    gender=gender_combobox.get()
    address=addressEntry.get(1.0,END)

    file=openpyxl.load_workbook("Backened_data.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)

    file.save(r"C:/Users/Ritz/IT_PROJECTS/PYTHON/DATA ENTRY FROM USER/Backened_data.xlsx")

    messagebox.showinfo('info','Detail added!!!')
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)
    

def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)


###########icon
icon_image=PhotoImage(file="logo1.png")
root.iconphoto(False,icon_image)


#######heading
Label(root,text="Please fill out this Entry form:", font="arial 15 bold",bg="lightblue",fg="black").place(x=20,y=30)


#######label
Label(root,text="Name", font=23,bg="lightblue",fg="black").place(x=50,y=100)
Label(root,text="Contact No.", font=23,bg="lightblue",fg="black").place(x=50,y=150)
Label(root,text="Age", font=23,bg="lightblue",fg="black").place(x=50,y=200)
Label(root,text="Gender", font=23,bg="lightblue",fg="black").place(x=370,y=200)
Label(root,text="Address", font=23,bg="lightblue",fg="black").place(x=50,y=250)


############Entry
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

######gender
gender_combobox=Combobox(root,values=['Male','Female'],font='arial 14',state='r',width=14)
gender_combobox.place(x=524,y=200)
gender_combobox.set('Male')

addressEntry=Text(root,width=45,height=4,bd=2,bg="lightgray",font=14)
addressEntry.place(x=198,y=250)

nameEntry = Entry(root,textvariable=nameValue,width=45,bd=2,font=20,bg="lightgray").place(x=200,y=100)
contactEntry = Entry(root,textvariable=contactValue,width=45,bd=2,font=20,bg="lightgray").place(x=200,y=150)
ageEntry = Entry(root,textvariable=AgeValue,width=13,bd=2,font=20,bg="lightgray").place(x=199,y=200)


Button(root,text="Submit",bg="lightblue",bd=4,fg="black",width=13,height=2,font=25,command=submit).place(x=198,y=360)
Button(root,text="Clear",bg="lightblue",bd=4,fg="black",width=13,height=2,font=25,command=clear).place(x=370,y=360)
Button(root,text="Exit",bg="lightblue",bd=4,fg="black",width=13,height=2,font=25,command=lambda:root.destroy()).place(x=542,y=360)























root.mainloop()
